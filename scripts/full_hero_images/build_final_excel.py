#!/usr/bin/env python3
"""Build the final full-field Excel workbook with embedded real-hardware images."""
from __future__ import annotations

import argparse
import csv
import json
import zipfile
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


MAIN_COLUMNS = [
    ("project_id", "项目ID"),
    ("__hero_image__", "英雄图"),
    ("hero_image_source_url", "英雄图来源"),
    ("hero_image_status", "图片核验状态"),
    ("name", "名称"),
    ("platform", "平台"),
    ("url", "原始链接"),
    ("source_domain", "来源域名"),
    ("thumbnail_url", "原始缩略图URL（仅追溯）"),
    ("thumbnail_type", "原始缩略图类型"),
    ("published_date", "发布时间"),
    ("updated_date", "最近更新"),
    ("description_zh", "项目简介（中文）"),
    ("keywords_zh", "核心关键词（中文）"),
    ("category", "类别"),
    ("stars_or_support", "支持量/Star"),
    ("hardware_license", "硬件许可"),
    ("software_license", "软件许可"),
    ("open_source_completeness", "开源完整度"),
    ("market_validation", "市场验证"),
    ("typical_competitors", "典型竞品"),
    ("commercial_value", "商业价值"),
    ("improvement_direction", "待改进方向"),
    ("target_customer", "目标客户"),
    ("suggested_price_low_cny", "建议售价下限（元）"),
    ("suggested_price_high_cny", "建议售价上限（元）"),
    ("manufacturing_difficulty", "量产难度"),
    ("after_sales_risk", "售后风险"),
    ("compliance_risk", "合规风险"),
    ("raw_commercial_score", "原始商业评分"),
    ("normalized_commercial_score", "正态化商业评分"),
    ("score_reason", "评分理由"),
    ("data_quality", "数据质量"),
    ("review_status", "评审状态"),
    ("crawl_time", "抓取时间"),
    ("description", "原文描述（追溯）"),
    ("keywords", "原始关键词（追溯）"),
]

TECH_COLUMNS = [
    "project_id", "hero_image_source_url", "hero_image_local_filename", "hero_image_status",
    "hero_image_reason", "candidate_count", "image_processed_at",
    "description_translation_status", "keywords_translation_status",
]

WIDTHS = [
    18, 24, 34, 22, 25, 13, 34, 18, 32, 17, 13, 13, 44, 32, 18, 13, 16, 16,
    26, 25, 28, 32, 32, 26, 15, 15, 13, 13, 13, 16, 17, 52, 12, 18, 20, 44, 32,
]

NUMERIC_FIELDS = {
    "stars_or_support", "suggested_price_low_cny", "suggested_price_high_cny",
    "manufacturing_difficulty", "after_sales_risk", "compliance_risk",
    "raw_commercial_score", "normalized_commercial_score",
}


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def number_or_text(value: str) -> Any:
    if value is None or value == "":
        return ""
    try:
        number = float(value)
        return int(number) if number.is_integer() else number
    except (TypeError, ValueError):
        return value


def add_hyperlink(cell, url: str, label: str) -> None:
    if not url:
        return
    cell.value = label
    cell.hyperlink = url
    cell.style = "Hyperlink"


def apply_header(ws, row: int, end_col: int) -> None:
    fill = PatternFill("solid", fgColor="245A87")
    font = Font(bold=True, color="FFFFFF", size=10)
    border = Border(bottom=Side(style="thin", color="AAB8C8"))
    for col in range(1, end_col + 1):
        cell = ws.cell(row, col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def build_workbook(input_dir: Path, output_path: Path) -> dict[str, Any]:
    csv_path = input_dir / "hardware_opportunities_full_hero_images.csv"
    images_dir = input_dir / "hero_images"
    summary_path = input_dir / "summary.json"
    source_status_path = input_dir / "source_status.csv"
    scoring_path = input_dir / "SCORING_METHOD.md"

    rows = read_csv(csv_path)
    if len(rows) != 10_500:
        raise RuntimeError(f"Expected 10500 rows, got {len(rows)}")
    if len({row["project_id"] for row in rows}) != 10_500:
        raise RuntimeError("project_id is not unique")
    summary = json.loads(summary_path.read_text(encoding="utf-8"))

    wb = Workbook()
    ws = wb.active
    ws.title = "项目机会库"
    summary_ws = wb.create_sheet("摘要")
    source_ws = wb.create_sheet("来源状态")
    scoring_ws = wb.create_sheet("评分方法")
    tech_ws = wb.create_sheet("处理状态")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(MAIN_COLUMNS))
    ws.cell(1, 1, "开源硬件商业化机会库：10,500条全字段真实英雄图版")
    ws.cell(1, 1).fill = PatternFill("solid", fgColor="17365D")
    ws.cell(1, 1).font = Font(bold=True, color="FFFFFF", size=18)
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(MAIN_COLUMNS))
    ws.cell(2, 1,
        "保留原始32个商业/评审字段；新增真实硬件英雄图、图片来源和核验状态；"
        "项目简介与关键词提供中文字段，同时保留英文原文。没有高置信度实物图的项目保持空白。"
    )
    ws.cell(2, 1).fill = PatternFill("solid", fgColor="EAF2F8")
    ws.cell(2, 1).font = Font(color="1F3A56", size=11)
    ws.cell(2, 1).alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 42

    for col, (_, label) in enumerate(MAIN_COLUMNS, start=1):
        ws.cell(3, col, label)
    apply_header(ws, 3, len(MAIN_COLUMNS))
    ws.row_dimensions[3].height = 34

    thin = Side(style="thin", color="D5DEE9")
    body_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    image_count = 0
    status_counts: Counter[str] = Counter()

    for excel_row, row in enumerate(rows, start=4):
        for col, (key, _) in enumerate(MAIN_COLUMNS, start=1):
            cell = ws.cell(excel_row, col)
            if key == "__hero_image__":
                value: Any = ""
            elif key in NUMERIC_FIELDS:
                value = number_or_text(row.get(key, ""))
            else:
                value = row.get(key, "")
            cell.value = value
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = body_border

        add_hyperlink(ws.cell(excel_row, 3), row.get("hero_image_source_url", ""), "核对英雄图")
        add_hyperlink(ws.cell(excel_row, 7), row.get("url", ""), "打开项目")
        add_hyperlink(ws.cell(excel_row, 9), row.get("thumbnail_url", ""), "原始URL")

        image_name = row.get("hero_image_local_filename", "")
        image_path = images_dir / image_name if image_name else None
        if image_path and image_path.is_file():
            image = XLImage(str(image_path))
            image.width = 120
            image.height = 70
            ws.add_image(image, f"B{excel_row}")
            image_count += 1
        ws.row_dimensions[excel_row].height = 58
        status_counts[row.get("hero_image_status", "")] += 1

    for index, width in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "E4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(MAIN_COLUMNS))}{len(rows)+3}"
    ws.sheet_view.showGridLines = False
    ws.conditional_formatting.add(
        f"AD4:AE{len(rows)+3}",
        ColorScaleRule(start_type="min", start_color="F8696B", mid_type="percentile", mid_value=50,
                       mid_color="FFEB84", end_type="max", end_color="63BE7B"),
    )

    # Summary sheet.
    summary_ws["A1"] = "交付摘要"
    summary_ws["A1"].fill = PatternFill("solid", fgColor="17365D")
    summary_ws["A1"].font = Font(bold=True, color="FFFFFF", size=17)
    summary_ws.merge_cells("A1:D1")
    summary_rows = [
        ("总记录数", len(rows)),
        ("唯一项目ID", len({row['project_id'] for row in rows})),
        ("内嵌英雄图", image_count),
        ("英雄图覆盖率", image_count / len(rows)),
        ("图像结果覆盖", summary.get("image_result_rows")),
        ("翻译结果覆盖", summary.get("translation_rows")),
        ("生成时间", summary.get("generated_at")),
    ]
    for r, (label, value) in enumerate(summary_rows, start=3):
        summary_ws.cell(r, 1, label)
        summary_ws.cell(r, 2, value)
    summary_ws["B6"].number_format = "0.00%"
    summary_ws.column_dimensions["A"].width = 24
    summary_ws.column_dimensions["B"].width = 28
    summary_ws["D3"] = "图片核验状态"
    summary_ws["E3"] = "数量"
    for i, (status, count) in enumerate(status_counts.most_common(), start=4):
        summary_ws.cell(i, 4, status)
        summary_ws.cell(i, 5, count)
    summary_ws.column_dimensions["D"].width = 34
    summary_ws.column_dimensions["E"].width = 12

    # Source status sheet.
    source_rows = read_csv(source_status_path) if source_status_path.exists() else []
    if source_rows:
        fields = list(source_rows[0].keys())
        source_ws.append(fields)
        apply_header(source_ws, 1, len(fields))
        for row in source_rows:
            source_ws.append([row.get(field, "") for field in fields])
        source_ws.freeze_panes = "A2"
        source_ws.auto_filter.ref = source_ws.dimensions
        for col in range(1, len(fields)+1):
            source_ws.column_dimensions[get_column_letter(col)].width = 22

    # Scoring method sheet.
    scoring_text = scoring_path.read_text(encoding="utf-8") if scoring_path.exists() else ""
    scoring_ws["A1"] = "评分方法"
    scoring_ws["A1"].fill = PatternFill("solid", fgColor="17365D")
    scoring_ws["A1"].font = Font(bold=True, color="FFFFFF", size=17)
    scoring_ws.column_dimensions["A"].width = 120
    for r, line in enumerate(scoring_text.splitlines(), start=3):
        scoring_ws.cell(r, 1, line)
        scoring_ws.cell(r, 1).alignment = Alignment(wrap_text=True, vertical="top")

    # Technical status sheet.
    tech_ws.append(TECH_COLUMNS)
    apply_header(tech_ws, 1, len(TECH_COLUMNS))
    for row in rows:
        tech_ws.append([row.get(field, "") for field in TECH_COLUMNS])
    tech_ws.freeze_panes = "A2"
    tech_ws.auto_filter.ref = tech_ws.dimensions
    for col in range(1, len(TECH_COLUMNS)+1):
        tech_ws.column_dimensions[get_column_letter(col)].width = 28

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    # Reopen and structural verification.
    check_wb = load_workbook(output_path, read_only=True, data_only=False)
    if check_wb["项目机会库"].max_row != 10_503:
        raise RuntimeError(f"Unexpected main sheet rows: {check_wb['项目机会库'].max_row}")
    required_sheets = {"项目机会库", "摘要", "来源状态", "评分方法", "处理状态"}
    if not required_sheets.issubset(set(check_wb.sheetnames)):
        raise RuntimeError(f"Missing sheets: {required_sheets - set(check_wb.sheetnames)}")
    check_wb.close()

    with zipfile.ZipFile(output_path) as archive:
        embedded = [name for name in archive.namelist() if name.startswith("xl/media/")]
    if len(embedded) != image_count:
        raise RuntimeError(f"Embedded image count mismatch: {len(embedded)} != {image_count}")

    return {
        "rows": len(rows),
        "unique_project_ids": len({row["project_id"] for row in rows}),
        "embedded_images": image_count,
        "xlsx_bytes": output_path.stat().st_size,
        "sheets": ["项目机会库", "摘要", "来源状态", "评分方法", "处理状态"],
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-dir", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()
    report = build_workbook(Path(args.input_dir), Path(args.output))
    report_path = Path(args.output).with_suffix(".validation.json")
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2), flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
